from openpyxl import Workbook
import io
import base64
from itertools import chain
from django.db.models.fields.files import (
    FieldFile, FileField, ImageFieldFile, ImageField
)

class Excel:

    def __init__(self, filename, extention, data, saved=False):
        self.saved = saved
        self.filename = filename
        self.extention = extention
        self._exclude_type = [FieldFile, FileField, ImageFieldFile, ImageField]
        self.data = self.serializer_list(data)
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        # self.params = {"instance": None, "fields": None, "exclude": None, "exclude_type": [FieldFile, FileField, ImageFieldFile, ImageField]} if params == {} else params

    def serializer_list(self, data):
        serialized = []
        for item in data:
            serialized.append(self.custom_model_to_dict(item, exclude_type=self._exclude_type))
        return serialized

    def set_columns_name(self):
        keys = sorted(self.data[0].keys())
        for column, key in enumerate(keys):
            cell = self.sheet.cell(row=1, column=column+1)  # columna+1 porque el índice de la primera columna es 1
            cell.value = key
    
    def write_rows(self):
        keys = sorted(self.data[0].keys())
        for column, key in enumerate(keys):
            row = 2
            for obj in self.data:
                cell = self.sheet.cell(row=row, column=column+1)  # columna+1 porque el índice de la primera columna es 1
                cell.value = obj[key]
                row += 1

    def export(self):
        self.set_columns_name()
        self.write_rows()
        if self.saved:
            self.__save()
    
    def save(self):
        output = io.BytesIO()
        self.workbook.save(output)
        return base64.b64encode(output.getvalue()).decode()


    def __save(self):
        self.workbook.save(f"{self.filename}.{self.extention}")


    def custom_model_to_dict(self, instance, fields=None, exclude=None, exclude_type: list=None):
        
        opts = instance._meta
        data = {}
        for f in chain(opts.concrete_fields, opts.private_fields, opts.many_to_many):
            if not getattr(f, "editable", False):
                continue
            if fields is not None and f.name not in fields:
                continue
            if exclude and f.name in exclude:
                continue

            # Exclude fields file
            if exclude_type is not None:
                if any([isinstance(f, _type) for _type in exclude_type]):
                    continue

            data[f.name] = f.value_from_object(instance)
        return data



